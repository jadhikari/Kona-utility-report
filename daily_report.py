from selenium import webdriver
import chromedriver_autoinstaller
import time
import re
import openpyxl
from selenium.webdriver.common.by import By

chromedriver_autoinstaller.install()  # Check if the current version of chromedriver exists
# and if it doesn't exist, download it automatically,
# then add chromedriver to path

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.kireilife.net/")
time.sleep(2)

#there are 4 user, you can chenge it according to your requerement.
user_id = "hecj5784"
password = "BrawnKona2018"

id_name = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[2]/div/ul/li[5]/div[2]/form/div[2]/input')
id_name.send_keys(user_id)
id_pass = driver.find_element(By.XPATH,
                              '/html/body/div[1]/div/div[1]/div/div[2]/div/ul/li[5]/div[2]/form/div[2]/div[1]/input')
id_pass.send_keys(password)

log_in = driver.find_element(By.XPATH,
                             value='/html/body/div[1]/div/div[1]/div/div[2]/div/ul/li[5]/div[2]/form/div[2]/button')
log_in.click()
time.sleep(10)

driver.find_element(By.XPATH, value='/html/body/div[1]/div/div[1]/div/div[8]/div/ul/li[2]/a/img').click()
driver.find_element(By.XPATH, value='/html/body/div[1]/div/div[1]/div/div[8]/div/table[2]/tbody/tr[2]/td[2]/a').click()
time.sleep(3)

driver.find_element(By.XPATH, value='/html/body/div/div[2]/div[1]/button').click()
time.sleep(2)

s_number = []
f_data = []
s_data = []
dates = []
class MyClass:
    # Define a method inside the class
    def my_method(self):
        time.sleep(1)
        iframe = driver.find_element(By.XPATH, '//*[@id="mieruframe"]')
        driver.switch_to.frame(iframe)
        e = driver.find_element(By.XPATH, value='//*[@id="mi_title"]').text
        print(e)
        table = driver.find_element(By.XPATH, value='//*[@id="mi_month_list_table"]/tbody')
        # Extract data from the first and second columns

        first_column_data = []
        second_column_data = []
        # Iterate through rows and extract data from the first and second columns
        rows = table.find_elements(By.TAG_NAME, 'tr')
        for row in rows:
            columns = row.find_elements(By.TAG_NAME, 'td')
            if len(columns) >= 2:
                # Extract data from the first and second columns
                input_string = columns[0].text
                formatted_date = re.sub(r"(\d{2})月(\d{2})日", r"\1/\2", input_string)
                first_column_data.append(formatted_date)
                second_column_data.append(columns[1].text)
        f_data.append(first_column_data)
        s_data.append(second_column_data)

        # Print the extracted data


my_object = MyClass()

#change the first number in () according to the User ID. You can find the num in read_me
for x in range(100,0,-1):
    path = '//*[@id="contractInfo'
    path += str(x)
    path += ' "]/table/tbody/tr[4]/td[1]'
    print(x)
    td = driver.find_element(By.XPATH, value= path)
    number = td.text
    if "/" in number:
        button = '//*[@id="contractInfo'
        button += str(x)
        button += ' "]/table/tbody'
        time.sleep(1)
        element = driver.find_element(By.XPATH, value=button)
        driver.execute_script("arguments[0].click();", element)
        time.sleep(4)
    else:
        s_number.append(number)
        button = '//*[@id="contractInfo'
        button += str(x)
        button += ' "]/table/tbody//button'
        element = driver.find_element(By.XPATH, value=button)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", element)
        time.sleep(3)
        driver.find_element(By.XPATH, value='//*[@id="menu_day"]').click()
        time.sleep(4)
        iframe = driver.find_element(By.XPATH, '//*[@id="mieruframe"]')
        driver.switch_to.frame(iframe)
        d = driver.find_element(By.XPATH, value='//*[@id="mi_title"]').text
        # change the date according to your need
        if '2023年09月分' in d:
            time.sleep(2)

            driver.switch_to.default_content()
            my_object.my_method()

        else:
            driver.find_element(By.XPATH, value='//*[@id="before_btn"]/img').click()
            driver.switch_to.default_content()
            time.sleep(3)
            my_object.my_method()

    driver.switch_to.default_content()
    driver.find_element(By.XPATH, value='//*[@id="contListBtn"]').click()
    time.sleep(2)



print("First Column Data:", f_data)
print("Second Column Data:", s_data)
workbook = openpyxl.Workbook()
# Iterate through the values in array 'a' and create sheets with corresponding sublists
for idx, sheet_number in enumerate(s_number, start=1):
    sheet = workbook.create_sheet(title=sheet_number)  # Use 'Sheet_' prefix for sheet names
    data_b = f_data[idx - 1]  # Get the corresponding sublist from list 'b'
    data_c = s_data[idx - 1]  # Get the corresponding sublist from list 'c'

    # Populate the sheet with values from the corresponding sublists vertically
    for row_idx, (value_b, value_c) in enumerate(zip(data_b, data_c), start=1):
        sheet.cell(row=row_idx, column=1, value=value_b)  # Add sublist from 'b' vertically
        sheet.cell(row=row_idx, column=2, value=value_c)  # Add sublist from 'c' vertically

# Remove the default sheet created and save the workbook to a file
workbook.remove(workbook.active)
workbook.save('result.xlsx')
